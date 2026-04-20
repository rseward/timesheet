import datetime
from typing import Optional, List
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from pydantic import BaseModel

import pprint
import sqlalchemy
import base64
import io
import bluestone.timesheet.config as cfg
from bluestone.timesheet.jsonmodels import (
    ProjectJson,
    ReportRowJson,
    TimePeriodReportRequest,
    ClientPeriodReportRequest,
    TimekeeperPeriodReportRequest,
)

from bluestone.timesheet.data.daos import getDaoFactory
from bluestone.timesheet.data.models import BillingEvent, Project, Task, Timekeeper, Client
from bluestone.timesheet.auth.auth_bearer import JWTBearer, JWT_SECRET_KEY_ALGORITHMS

daos = getDaoFactory()

router = APIRouter(
    prefix="/api/reports",
    tags=["reports"],
    responses={404: {"description": "Not found"}},
)


def _calculate_hours(start_time: datetime.datetime, end_time: datetime.datetime) -> float:
    """Calculate hours between two datetimes, rounded to 1 decimal."""
    delta = end_time - start_time
    return round(delta.total_seconds() / 3600, 1)


def _build_time_period_rows(start_date: datetime.datetime, end_date: datetime.datetime) -> List[dict]:
    """Build report rows for Time Period Report.

    Returns all active billing events in the date range, enriched with
    client name, resource name, project title, hours, and bill rate.
    """
    session = daos.getBillingEventDao().getSession()

    q = session.query(
        BillingEvent, Project.title, Task.name, Timekeeper.first_name,
        Timekeeper.last_name, Timekeeper.bill_rate, Client.organisation
    )
    q = q.filter(BillingEvent.project_id == Project.project_id)
    q = q.filter(BillingEvent.task_id == Task.task_id)
    q = q.filter(BillingEvent.timekeeper_id == Timekeeper.timekeeper_id)
    q = q.filter(Project.client_id == Client.client_id)
    q = q.filter(BillingEvent.start_time >= start_date)
    q = q.filter(BillingEvent.start_time <= end_date)
    q = q.filter(BillingEvent.active)
    q = q.order_by(Client.organisation, Project.title, Timekeeper.last_name, BillingEvent.start_time)

    results = q.all()
    rows = []
    for row in results:
        event = row[0]
        project_title = row[1]
        task_name = row[2]
        first_name = row[3]
        last_name = row[4]
        bill_rate = row[5]
        client_org = row[6]

        hours = _calculate_hours(event.start_time, event.end_time)
        date_str = event.start_time.date().isoformat() if hasattr(event.start_time, 'date') else str(event.start_time)[:10]

        rows.append({
            "client": client_org,
            "resource": f"{first_name} {last_name}",
            "date": date_str,
            "hours": hours,
            "bill_rate": bill_rate,
            "task": event.log_message or task_name,
            "project": project_title,
        })

    return rows


def _build_client_period_rows(start_date: datetime.datetime, end_date: datetime.datetime,
                               client_id: int, project_id: Optional[int] = None) -> List[dict]:
    """Build report rows for Client Period Report.

    Returns active billing events for a specific client (and optionally project)
    in the date range.
    """
    session = daos.getBillingEventDao().getSession()

    q = session.query(
        BillingEvent, Project.title, Task.name, Timekeeper.first_name,
        Timekeeper.last_name, Timekeeper.bill_rate, Client.organisation
    )
    q = q.filter(BillingEvent.project_id == Project.project_id)
    q = q.filter(BillingEvent.task_id == Task.task_id)
    q = q.filter(BillingEvent.timekeeper_id == Timekeeper.timekeeper_id)
    q = q.filter(Project.client_id == Client.client_id)
    q = q.filter(Client.client_id == client_id)
    q = q.filter(BillingEvent.start_time >= start_date)
    q = q.filter(BillingEvent.start_time <= end_date)
    q = q.filter(BillingEvent.active)

    if project_id is not None:
        q = q.filter(BillingEvent.project_id == project_id)

    q = q.order_by(Project.title, Timekeeper.last_name, BillingEvent.start_time)

    results = q.all()
    rows = []
    for row in results:
        event = row[0]
        project_title = row[1]
        task_name = row[2]
        first_name = row[3]
        last_name = row[4]
        bill_rate = row[5]
        client_org = row[6]

        hours = _calculate_hours(event.start_time, event.end_time)
        date_str = event.start_time.date().isoformat() if hasattr(event.start_time, 'date') else str(event.start_time)[:10]

        rows.append({
            "client": client_org,
            "resource": f"{first_name} {last_name}",
            "date": date_str,
            "hours": hours,
            "bill_rate": bill_rate,
            "task": event.log_message or task_name,
            "project": project_title,
        })

    return rows


def _build_timekeeper_period_rows(start_date: datetime.datetime, end_date: datetime.datetime,
                                   timekeeper_id: int) -> List[dict]:
    """Build report rows for TimeKeeper Period Report.

    Returns active billing events for a specific timekeeper in the date range.
    """
    session = daos.getBillingEventDao().getSession()

    q = session.query(
        BillingEvent, Project.title, Task.name, Timekeeper.first_name,
        Timekeeper.last_name, Timekeeper.bill_rate, Client.organisation
    )
    q = q.filter(BillingEvent.project_id == Project.project_id)
    q = q.filter(BillingEvent.task_id == Task.task_id)
    q = q.filter(BillingEvent.timekeeper_id == Timekeeper.timekeeper_id)
    q = q.filter(Project.client_id == Client.client_id)
    q = q.filter(BillingEvent.timekeeper_id == timekeeper_id)
    q = q.filter(BillingEvent.start_time >= start_date)
    q = q.filter(BillingEvent.start_time <= end_date)
    q = q.filter(BillingEvent.active)
    q = q.order_by(BillingEvent.start_time)

    results = q.all()
    rows = []
    for row in results:
        event = row[0]
        project_title = row[1]
        task_name = row[2]
        first_name = row[3]
        last_name = row[4]
        bill_rate = row[5]
        client_org = row[6]

        hours = _calculate_hours(event.start_time, event.end_time)
        date_str = event.start_time.date().isoformat() if hasattr(event.start_time, 'date') else str(event.start_time)[:10]

        rows.append({
            "client": client_org,
            "resource": f"{first_name} {last_name}",
            "date": date_str,
            "hours": hours,
            "bill_rate": bill_rate,
            "task": event.log_message or task_name,
            "project": project_title,
        })

    return rows


def _compute_summary(rows: List[dict]) -> dict:
    """Compute summary statistics from report rows."""
    if not rows:
        return {
            "total_hours": 0.0,
            "total_rows": 0,
            "unique_resources": 0,
            "unique_projects": 0,
            "unique_clients": 0,
            "date_range": {"start": None, "end": None},
        }

    total_hours = sum(r["hours"] for r in rows)
    dates = [r["date"] for r in rows]

    return {
        "total_hours": round(total_hours, 1),
        "total_rows": len(rows),
        "unique_resources": len(set(r["resource"] for r in rows)),
        "unique_projects": len(set(r["project"] for r in rows if r.get("project"))),
        "unique_clients": len(set(r["client"] for r in rows if r.get("client"))),
        "date_range": {
            "start": min(dates),
            "end": max(dates),
        },
    }


@router.get(
    "/",
    response_model=dict,
    dependencies=[Depends(JWTBearer())],
)
def index() -> dict:
    """List available report types."""
    return {
        "reports": [
            {
                "type": "client-period",
                "name": "Client Period Report",
                "description": "Generate reports for specific clients within a date range",
                "endpoint": "/api/reports/client-period",
                "params": ["start_date", "end_date", "client_id", "project_id (optional)"],
            },
            {
                "type": "timekeeper-period",
                "name": "TimeKeeper Period Report",
                "description": "Comprehensive time tracking report for a specific timekeeper",
                "endpoint": "/api/reports/timekeeper-period",
                "params": ["start_date", "end_date", "timekeeper_id"],
            },
            {
                "type": "time-period",
                "name": "Time Period Report",
                "description": "General time analysis across all clients and resources",
                "endpoint": "/api/reports/time-period",
                "params": ["start_date", "end_date"],
            },
        ]
    }


@router.get(
    "/time-period",
    response_model=dict,
    dependencies=[Depends(JWTBearer())],
)
def time_period_report(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
) -> dict:
    """Generate a Time Period Report.

    Returns all billing events within the specified date range,
    including client, resource, date, hours, bill rate, task, and project.
    Also returns summary statistics.
    """
    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if sd > ed:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))

    rows = _build_time_period_rows(start_dt, end_dt)
    summary = _compute_summary(rows)

    return {
        "report_type": "time-period",
        "start_date": start_date,
        "end_date": end_date,
        "summary": summary,
        "rows": rows,
    }


@router.get(
    "/client-period",
    response_model=dict,
    dependencies=[Depends(JWTBearer())],
)
def client_period_report(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    client_id: int = Query(..., description="Client ID"),
    project_id: Optional[int] = Query(None, description="Project ID (optional filter)"),
) -> dict:
    """Generate a Client Period Report.

    Returns billing events for a specific client within the date range.
    Optionally filter by project_id.
    """
    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if sd > ed:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))

    rows = _build_client_period_rows(start_dt, end_dt, client_id, project_id)
    summary = _compute_summary(rows)

    return {
        "report_type": "client-period",
        "start_date": start_date,
        "end_date": end_date,
        "client_id": client_id,
        "project_id": project_id,
        "summary": summary,
        "rows": rows,
    }


@router.get(
    "/timekeeper-period",
    response_model=dict,
    dependencies=[Depends(JWTBearer())],
)
def timekeeper_period_report(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    timekeeper_id: int = Query(..., description="Timekeeper ID"),
) -> dict:
    """Generate a TimeKeeper Period Report.

    Returns billing events for a specific timekeeper within the date range.
    """
    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if sd > ed:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))

    rows = _build_timekeeper_period_rows(start_dt, end_dt, timekeeper_id)
    summary = _compute_summary(rows)

    return {
        "report_type": "timekeeper-period",
        "start_date": start_date,
        "end_date": end_date,
        "timekeeper_id": timekeeper_id,
        "summary": summary,
        "rows": rows,
    }


@router.get(
    "/time-period/csv",
    dependencies=[Depends(JWTBearer())],
)
def time_period_report_csv(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
):
    """Generate a Time Period Report as CSV download."""
    from fastapi.responses import StreamingResponse
    import io
    import csv

    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))

    rows = _build_time_period_rows(start_dt, end_dt)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Client", "Resource", "Date", "Hours", "Billing Rate", "Task", "Project"])
    for row in rows:
        writer.writerow([
            row["client"] or "",
            row["resource"],
            row["date"],
            row["hours"],
            row["bill_rate"] or "",
            row["task"] or "",
            row["project"] or "",
        ])

    output.seek(0)
    filename = f"time-period-report-{start_date}-to-{end_date}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/client-period/csv",
    dependencies=[Depends(JWTBearer())],
)
def client_period_report_csv(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    client_id: int = Query(..., description="Client ID"),
    project_id: Optional[int] = Query(None, description="Project ID (optional)"),
):
    """Generate a Client Period Report as CSV download."""
    from fastapi.responses import StreamingResponse
    import io
    import csv

    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))

    rows = _build_client_period_rows(start_dt, end_dt, client_id, project_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Client", "Resource", "Date", "Hours", "Billing Rate", "Task", "Project"])
    for row in rows:
        writer.writerow([
            row["client"] or "",
            row["resource"],
            row["date"],
            row["hours"],
            row["bill_rate"] or "",
            row["task"] or "",
            row["project"] or "",
        ])

    output.seek(0)
    filename = f"client-period-report-{start_date}-to-{end_date}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/timekeeper-period/csv",
    dependencies=[Depends(JWTBearer())],
)
def timekeeper_period_report_csv(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    timekeeper_id: int = Query(..., description="Timekeeper ID"),
):
    """Generate a TimeKeeper Period Report as CSV download."""
    from fastapi.responses import StreamingResponse
    import io
    import csv

    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))

    rows = _build_timekeeper_period_rows(start_dt, end_dt, timekeeper_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Client", "Resource", "Date", "Hours", "Billing Rate", "Task", "Project"])
    for row in rows:
        writer.writerow([
            row["client"] or "",
            row["resource"],
            row["date"],
            row["hours"],
            row["bill_rate"] or "",
            row["task"] or "",
            row["project"] or "",
        ])

    output.seek(0)
    filename = f"timekeeper-period-report-{start_date}-to-{end_date}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/clients",
    dependencies=[Depends(JWTBearer())],
)
def list_clients():
    """List all active clients for report parameter selection."""
    ClientDao = daos.getClientDao()
    clients = ClientDao.getAll(include_inactive=False)
    return {
        "clients": [
            {"client_id": c.client_id, "organisation": c.organisation}
            for c in clients
        ]
    }


@router.get(
    "/timekeepers",
    dependencies=[Depends(JWTBearer())],
)
def list_timekeepers():
    """List all timekeepers for report parameter selection."""
    TimekeeperDao = daos.getTimekeeperDao()
    timekeepers = TimekeeperDao.getAll()
    return {
        "timekeepers": [
            {
                "timekeeper_id": t.timekeeper_id,
                "name": f"{t.first_name} {t.last_name}",
                "username": t.username,
            }
            for t in timekeepers
        ]
    }


@router.get(
    "/projects",
    dependencies=[Depends(JWTBearer())],
)
def list_projects(client_id: Optional[int] = Query(None, description="Filter by client ID")):
    """List all active projects for report parameter selection. Optionally filter by client_id."""
    ProjectDao = daos.getProjectDao()
    results = ProjectDao.getAll(include_inactive=False)

    projects = []
    for row in results:
        project = row[0]
        org = row[1]
        if client_id is not None and project.client_id != client_id:
            continue
        projects.append({
            "project_id": project.project_id,
            "title": project.title,
            "client_id": project.client_id,
            "client_name": org,
        })

    return {"projects": projects}


# ═══════════════════════════════════════════════════════════════════
# Excel Export Helpers
# ═══════════════════════════════════════════════════════════════════

EXCEL_COLUMNS = ["Client", "Resource", "Date", "Hours", "Billing Rate", "Task", "Project"]
EXCEL_KEYS = ["client", "resource", "date", "hours", "bill_rate", "task", "project"]


def _rows_to_xlsx(rows: List[dict], template_data: Optional[bytes] = None) -> bytes:
    """Convert report rows to an Excel (.xlsx) byte stream.

    If template_data is provided, the unformatted spreadsheet columns must
    match the template's structure. The function loads the template workbook,
    finds the first sheet, writes data starting at row 2 (row 1 = header),
    and returns the filled-in workbook.

    Without a template, a plain unformatted spreadsheet is produced.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    if template_data:
        wb = load_workbook(io.BytesIO(template_data))
        ws = wb.active
        # Write data rows starting at row 2 (row 1 is the template header)
        for i, row in enumerate(rows, start=2):
            for j, key in enumerate(EXCEL_KEYS, start=1):
                val = row.get(key, "")
                if val is None:
                    val = ""
                ws.cell(row=i, column=j, value=val)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        # Header row
        header_font = Font(bold=True)
        for j, col_name in enumerate(EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=j, value=col_name)
            cell.font = header_font
        # Data rows
        for i, row in enumerate(rows, start=2):
            for j, key in enumerate(EXCEL_KEYS, start=1):
                val = row.get(key, "")
                if val is None:
                    val = ""
                ws.cell(row=i, column=j, value=val)
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _excel_response(xlsx_bytes: bytes, filename: str):
    """Return a StreamingResponse for an Excel file."""
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ═══════════════════════════════════════════════════════════════════
# Excel Export Endpoints
# ═══════════════════════════════════════════════════════════════════

@router.get(
    "/time-period/excel",
    dependencies=[Depends(JWTBearer())],
)
def time_period_report_excel(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    template_id: Optional[int] = Query(None, description="Report template ID for formatted export"),
):
    """Generate a Time Period Report as Excel (.xlsx) download.
    If template_id is provided, the data is written into the template."""
    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    if sd > ed:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))
    rows = _build_time_period_rows(start_dt, end_dt)

    template_data = None
    if template_id is not None:
        tmpl = daos.getReportTemplateDao().getById(template_id)
        if not tmpl or not tmpl.active:
            raise HTTPException(status_code=404, detail="Template not found or inactive")
        template_data = base64.b64decode(tmpl.file_data)

    xlsx_bytes = _rows_to_xlsx(rows, template_data)
    filename = f"time-period-report-{start_date}-to-{end_date}.xlsx"
    return _excel_response(xlsx_bytes, filename)


@router.get(
    "/client-period/excel",
    dependencies=[Depends(JWTBearer())],
)
def client_period_report_excel(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    client_id: int = Query(..., description="Client ID"),
    project_id: Optional[int] = Query(None, description="Project ID (optional)"),
    template_id: Optional[int] = Query(None, description="Report template ID for formatted export"),
):
    """Generate a Client Period Report as Excel (.xlsx) download."""
    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    if sd > ed:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))
    rows = _build_client_period_rows(start_dt, end_dt, client_id, project_id)

    template_data = None
    if template_id is not None:
        tmpl = daos.getReportTemplateDao().getById(template_id)
        if not tmpl or not tmpl.active:
            raise HTTPException(status_code=404, detail="Template not found or inactive")
        template_data = base64.b64decode(tmpl.file_data)

    xlsx_bytes = _rows_to_xlsx(rows, template_data)
    filename = f"client-period-report-{start_date}-to-{end_date}.xlsx"
    return _excel_response(xlsx_bytes, filename)


@router.get(
    "/timekeeper-period/excel",
    dependencies=[Depends(JWTBearer())],
)
def timekeeper_period_report_excel(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    timekeeper_id: int = Query(..., description="Timekeeper ID"),
    template_id: Optional[int] = Query(None, description="Report template ID for formatted export"),
):
    """Generate a TimeKeeper Period Report as Excel (.xlsx) download."""
    try:
        sd = datetime.date.fromisoformat(start_date)
        ed = datetime.date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    if sd > ed:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date.")

    start_dt = datetime.datetime.combine(sd, datetime.time(0, 0))
    end_dt = datetime.datetime.combine(ed, datetime.time(23, 59, 59))
    rows = _build_timekeeper_period_rows(start_dt, end_dt, timekeeper_id)

    template_data = None
    if template_id is not None:
        tmpl = daos.getReportTemplateDao().getById(template_id)
        if not tmpl or not tmpl.active:
            raise HTTPException(status_code=404, detail="Template not found or inactive")
        template_data = base64.b64decode(tmpl.file_data)

    xlsx_bytes = _rows_to_xlsx(rows, template_data)
    filename = f"timekeeper-period-report-{start_date}-to-{end_date}.xlsx"
    return _excel_response(xlsx_bytes, filename)


# ═══════════════════════════════════════════════════════════════════
# Report Template Management
# ═══════════════════════════════════════════════════════════════════

@router.get(
    "/templates",
    dependencies=[Depends(JWTBearer())],
)
def list_report_templates(
    report_type: Optional[str] = Query(None, description="Filter by report type"),
):
    """List available report templates. Optionally filter by report_type."""
    dao = daos.getReportTemplateDao()
    templates = dao.getAll(report_type=report_type)
    return {
        "templates": [dao.toDict(t) for t in templates]
    }


@router.post(
    "/templates",
    dependencies=[Depends(JWTBearer())],
)
def upload_report_template(
    name: str = Form(..., description="Template name"),
    report_type: str = Form(..., description="Report type: client-period, timekeeper-period, or time-period"),
    file: UploadFile = File(..., description="Excel template file (.xlsx)"),
    description: Optional[str] = Form(None, description="Template description"),
    created_by: Optional[str] = Form(None, description="Uploader name/email"),
):
    """Upload a new report template.

    The template must be an .xlsx file with the same column structure as
    the unformatted export (Client, Resource, Date, Hours, Billing Rate,
    Task, Project). Row 1 is treated as the header row; data will be
    written starting at row 2.
    """
    if report_type not in ("client-period", "timekeeper-period", "time-period"):
        raise HTTPException(status_code=400, detail="Invalid report_type. Must be client-period, timekeeper-period, or time-period.")

    content = file.file.read()
    # Validate it's a valid xlsx
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        # Ensure at least one sheet
        if not wb.sheetnames:
            raise ValueError("No sheets in workbook")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    from bluestone.timesheet.data.models import ReportTemplate
    tmpl = ReportTemplate(
        name=name,
        description=description,
        report_type=report_type,
        filename=file.filename or "template.xlsx",
        file_data=base64.b64encode(content).decode("utf-8"),
        created_by=created_by,
        created_at=datetime.datetime.utcnow(),
        active=True,
    )
    dao = daos.getReportTemplateDao()
    dao.save(tmpl)
    dao.commit()

    return {"template_id": tmpl.template_id, "name": tmpl.name, "message": "Template uploaded successfully"}


@router.delete(
    "/templates/{template_id}",
    dependencies=[Depends(JWTBearer())],
)
def delete_report_template(template_id: int):
    """Delete a report template."""
    dao = daos.getReportTemplateDao()
    tmpl = dao.getById(template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    dao.delete(template_id)
    return {"message": f"Template {template_id} deleted"}